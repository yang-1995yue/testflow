"""
项目级测试用例管理API
提供项目下所有模块测试用例的聚合查询和批量操作
"""
from typing import Any, List, Optional
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, Form
from sqlalchemy import or_
from sqlalchemy.orm import Session, joinedload
from pydantic import BaseModel

from app.database import get_db
from app.models.user import User, ProjectMember, ProjectRole
from app.models.project import Project
from app.models.module import Module
from app.models.requirement import RequirementPoint
from app.models.testcase import TestPoint, TestCase, TestCaseStatus
from app.core.dependencies import get_current_active_user

router = APIRouter()


# ========== Pydantic 响应模型 ==========

class TestCaseItem(BaseModel):
    """测试用例扁平化展示"""
    id: int
    title: str
    description: Optional[str] = None
    preconditions: Optional[str] = None
    test_steps: Optional[List[dict]] = None
    expected_result: Optional[str] = None
    design_method: Optional[str] = None
    test_category: Optional[str] = None
    priority: Optional[str] = None
    status: Optional[str] = None
    module_id: Optional[int] = None
    module_name: Optional[str] = None
    test_point_id: Optional[int] = None
    test_point_content: Optional[str] = None

    class Config:
        from_attributes = True


class ModuleTestCasesGroup(BaseModel):
    """按模块分组的测试用例"""
    id: int
    name: str
    test_cases: List[TestCaseItem]

    class Config:
        from_attributes = True


class BatchDeleteRequest(BaseModel):
    """批量删除请求"""
    ids: List[int]


# ========== 权限检查 ==========

def check_project_access(project_id: int, user: User, db: Session) -> Project:
    """检查用户对项目的访问权限，返回项目对象"""
    from app.models.user import UserRole
    
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在")
    
    # 管理员或项目所有者
    if user.role == UserRole.ADMIN or project.owner_id == user.id:
        return project
    
    # 检查是否为项目成员
    member = db.query(ProjectMember).filter(
        ProjectMember.project_id == project_id,
        ProjectMember.user_id == user.id
    ).first()
    
    if not member:
        raise HTTPException(status_code=403, detail="无权访问此项目")
    
    return project


def check_project_edit_permission(project_id: int, user: User, db: Session) -> Project:
    """检查用户是否有编辑权限（成员或管理员）"""
    from app.models.user import UserRole
    
    project = check_project_access(project_id, user, db)
    
    # 管理员或项目所有者
    if user.role == UserRole.ADMIN or project.owner_id == user.id:
        return project
    
    # 检查是否为编辑角色
    member = db.query(ProjectMember).filter(
        ProjectMember.project_id == project_id,
        ProjectMember.user_id == user.id
    ).first()
    
    if member and member.role == ProjectRole.VIEWER:
        raise HTTPException(status_code=403, detail="查看者无编辑权限")
    
    return project


# ========== API 路由 ==========

@router.get("/projects/{project_id}/test-cases", response_model=List[Any])
async def get_project_test_cases(
    project_id: int,
    view_mode: str = Query("hierarchy", regex="^(hierarchy|flat)$"),
    keyword: Optional[str] = None,
    status: Optional[str] = None,
    priority: Optional[str] = None,
    test_category: Optional[str] = None,
    module_id: Optional[int] = None,
    design_method: Optional[str] = None,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    获取项目下所有模块的测试用例
    
    - view_mode: hierarchy (按模块分组) | flat (扁平列表)
    - keyword: 搜索标题/内容
    - status: 筛选状态
    - priority: 筛选优先级
    - test_category: 筛选测试分类
    - module_id: 筛选所属模块
    - design_method: 筛选设计方法
    """
    # 权限检查
    check_project_access(project_id, current_user, db)
    
    # 获取项目下所有模块
    modules = db.query(Module).filter(Module.project_id == project_id).all()
    module_ids = [m.id for m in modules]
    module_map = {m.id: m.name for m in modules}
    
    if not module_ids:
        return []
    
    # 获取所有模块的需求点
    requirement_points = db.query(RequirementPoint).filter(
        RequirementPoint.module_id.in_(module_ids)
    ).all()
    rp_ids = [rp.id for rp in requirement_points]
    rp_module_map = {rp.id: rp.module_id for rp in requirement_points}
    

    
    # 获取所有测试点
    test_points = db.query(TestPoint).filter(
        TestPoint.requirement_point_id.in_(rp_ids)
    ).all()
    tp_ids = [tp.id for tp in test_points]
    tp_rp_map = {tp.id: tp.requirement_point_id for tp in test_points}
    tp_content_map = {tp.id: tp.content for tp in test_points}
    

    
    # 构建测试用例查询
    # 构建测试用例查询
    query = db.query(TestCase).filter(
        or_(
            TestCase.test_point_id.in_(tp_ids),
            TestCase.module_id.in_(module_ids),
            TestCase.project_id == project_id
        )
    )
    
    # 应用筛选条件
    if keyword:
        query = query.filter(TestCase.title.ilike(f"%{keyword}%"))
    if status:
        query = query.filter(TestCase.status == status)
    if priority:
        query = query.filter(TestCase.priority == priority)
    if test_category:
        query = query.filter(TestCase.test_category == test_category)
    if design_method:
        query = query.filter(TestCase.design_method == design_method)
    if module_id:
        # 直接按模块ID筛选，支持两种方式：
        # 1. 直接关联的module_id
        # 2. 通过测试点-需求点关联的module_id
        query = query.filter(
            or_(
                TestCase.module_id == module_id,
                TestCase.test_point_id.in_(
                    db.query(TestPoint.id).filter(
                        TestPoint.requirement_point_id.in_(
                            db.query(RequirementPoint.id).filter(
                                RequirementPoint.module_id == module_id
                            ).subquery()
                        )
                    ).subquery()
                )
            )
        )
    
    test_cases = query.all()
    
    # 构建响应
    if view_mode == "flat":
        result = []
        for tc in test_cases:
            tp_id = tc.test_point_id
            rp_id = tp_rp_map.get(tp_id)
            module_id = rp_module_map.get(rp_id) if rp_id else None
            
            # 确定模块信息
            final_module_id = tc.module_id if tc.module_id else module_id
            final_module_name = "未分类"
            if final_module_id:
                final_module_name = module_map.get(final_module_id, "未分类")
            elif tc.import_module_name:
                final_module_name = tc.import_module_name

            result.append(TestCaseItem(
                id=tc.id,
                title=tc.title,
                description=tc.description,
                preconditions=tc.preconditions,
                test_steps=tc.test_steps,
                expected_result=tc.expected_result,
                design_method=tc.design_method,
                test_category=tc.test_category,
                priority=tc.priority,
                status=tc.status,
                module_id=final_module_id,
                module_name=final_module_name,
                test_point_id=tp_id,
                test_point_content=tp_content_map.get(tp_id)
            ).model_dump())
        return result
    
    # hierarchy 模式：按模块分组
    module_cases = {m.id: [] for m in modules}
    module_cases[0] = []  # 未分类
    
    for tc in test_cases:
        tp_id = tc.test_point_id
        rp_id = tp_rp_map.get(tp_id)
        
        # 确定模块归属
        # 1. 优先使用直接关联的 module_id
        # 2. 其次使用通过测试点关联的 module_id
        # 3. 如果都没有，则归为未分类 (0)
        module_id = tc.module_id if tc.module_id else (rp_module_map.get(rp_id) if rp_id else 0)
        
        # 确定模块名称显示
        module_name = "未分类"
        if module_id:
            module_name = module_map.get(module_id, "未分类")
        elif tc.import_module_name:
            # 如果是未分类但有导入时的模块名，显示该名称（但在分组时仍归为未分类）
            module_name = tc.import_module_name
        
        item = TestCaseItem(
            id=tc.id,
            title=tc.title,
            description=tc.description,
            preconditions=tc.preconditions,
            test_steps=tc.test_steps,
            expected_result=tc.expected_result,
            design_method=tc.design_method,
            test_category=tc.test_category,
            priority=tc.priority,
            status=tc.status,
            module_id=module_id,
            module_name=module_name,
            test_point_id=tp_id,
            test_point_content=tp_content_map.get(tp_id)
        )
        
        if module_id in module_cases:
            module_cases[module_id].append(item.model_dump())
        else:
            module_cases[0].append(item.model_dump())
    
    result = [
        ModuleTestCasesGroup(id=m.id, name=m.name, test_cases=module_cases.get(m.id, [])).model_dump()
        for m in modules
    ]
    
    # 添加未分类（如果有）
    if module_cases[0]:
        result.append(ModuleTestCasesGroup(id=0, name="未分类", test_cases=module_cases[0]).model_dump())
    
    return result


@router.delete("/projects/{project_id}/test-cases/batch")
async def batch_delete_test_cases(
    project_id: int,
    request: BatchDeleteRequest,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """批量删除测试用例"""
    # 权限检查（需要编辑权限）
    check_project_edit_permission(project_id, current_user, db)
    
    if not request.ids:
        raise HTTPException(status_code=400, detail="请选择要删除的用例")
    
    # 验证用例属于该项目
    modules = db.query(Module).filter(Module.project_id == project_id).all()
    module_ids = [m.id for m in modules]
    
    rp_ids = [rp.id for rp in db.query(RequirementPoint.id).filter(
        RequirementPoint.module_id.in_(module_ids)
    ).all()]
    
    tp_ids = [tp.id for tp in db.query(TestPoint.id).filter(
        TestPoint.requirement_point_id.in_(rp_ids)
    ).all()]
    
    # 执行删除
    deleted = db.query(TestCase).filter(
        TestCase.id.in_(request.ids),
        TestCase.test_point_id.in_(tp_ids)
    ).delete(synchronize_session=False)
    
    db.commit()
    
    return {"deleted_count": deleted, "message": f"成功删除 {deleted} 条用例"}


class TestCaseUpdateRequest(BaseModel):
    """测试用例更新请求"""
    title: Optional[str] = None
    description: Optional[str] = None
    preconditions: Optional[str] = None
    test_steps: Optional[List[dict]] = None
    expected_result: Optional[str] = None
    design_method: Optional[str] = None
    test_category: Optional[str] = None
    priority: Optional[str] = None
    status: Optional[str] = None
    module_id: Optional[int] = None


def verify_test_case_belongs_to_project(case_id: int, project_id: int, db: Session) -> TestCase:
    """验证测试用例属于指定项目，返回用例对象"""
    # 获取项目的所有模块
    modules = db.query(Module).filter(Module.project_id == project_id).all()
    module_ids = [m.id for m in modules]
    
    # 获取模块的需求点
    rp_ids = []
    if module_ids:
        rp_ids = [rp.id for rp in db.query(RequirementPoint.id).filter(
            RequirementPoint.module_id.in_(module_ids)
        ).all()]
    
    # 获取测试点
    tp_ids = []
    if rp_ids:
        tp_ids = [tp.id for tp in db.query(TestPoint.id).filter(
            TestPoint.requirement_point_id.in_(rp_ids)
        ).all()]
    
    # 查找测试用例（支持通过 test_point_id 或直接 module_id 或直接 project_id 验证）
    test_case = db.query(TestCase).filter(
        TestCase.id == case_id,
        or_(
            TestCase.test_point_id.in_(tp_ids) if tp_ids else False,
            TestCase.module_id.in_(module_ids) if module_ids else False,
            TestCase.project_id == project_id
        )
    ).first()
    
    if not test_case:
        raise HTTPException(status_code=404, detail="测试用例不存在或不属于此项目")
    
    return test_case


@router.put("/projects/{project_id}/test-cases/{case_id}")
async def update_test_case(
    project_id: int,
    case_id: int,
    request: TestCaseUpdateRequest,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """更新单个测试用例"""
    # 权限检查
    check_project_edit_permission(project_id, current_user, db)
    
    # 验证用例属于该项目
    test_case = verify_test_case_belongs_to_project(case_id, project_id, db)
    
    # 更新字段
    update_data = request.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        if value is not None:
            setattr(test_case, key, value)
    
    test_case.edited_by_user = True
    db.commit()
    db.refresh(test_case)
    
    return {"message": "更新成功", "id": test_case.id}


@router.delete("/projects/{project_id}/test-cases/{case_id}")
async def delete_single_test_case(
    project_id: int,
    case_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """删除单个测试用例"""
    # 权限检查
    check_project_edit_permission(project_id, current_user, db)
    
    # 验证用例属于该项目
    test_case = verify_test_case_belongs_to_project(case_id, project_id, db)
    
    # 执行删除
    db.delete(test_case)
    db.commit()
    
    return {"message": "删除成功", "id": case_id}


class ExportRequest(BaseModel):
    """导出请求"""
    ids: Optional[List[int]] = None  # 指定导出的用例ID，为空则导出全部
    format: str = "excel"  # 导出格式：excel, xmind


@router.post("/projects/{project_id}/test-cases/export")
async def export_test_cases(
    project_id: int,
    request: ExportRequest,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    导出测试用例
    
    支持格式：
    - excel: Excel表格（测试步骤和预期结果分列展示）
    - xmind: 思维导图（按模块-用例-步骤层级展示）
    Excel列格式：
    - 序号
    - 所属模块
    - 用例标题
    - 前置条件
    - 测试步骤（多步骤以1. 2. 3.格式换行展示）
    - 预期结果（对应步骤以1. 2. 3.格式换行展示）
    - 优先级
    - 测试分类
    - 设计方法
    - 状态
    """
    from fastapi.responses import StreamingResponse
    from io import BytesIO
    from datetime import datetime
    from urllib.parse import quote
    
    print(f"📥 收到导出请求: project_id={project_id}, request={request.dict()}")
    
    # 权限检查
    try:
        check_project_access(project_id, current_user, db)
        print(f"✅ 权限检查通过")
    except Exception as e:
        print(f"❌ 权限检查失败: {e}")
        raise
    
    # 获取项目信息
    project = db.query(Project).filter(Project.id == project_id).first()
    print(f"📄 获取到项目信息: {project.name}")
    
    # 获取项目下所有模块
    modules = db.query(Module).filter(Module.project_id == project_id).all()
    module_ids = [m.id for m in modules]
    module_map = {m.id: m.name for m in modules}
    
    if not module_ids:
        raise HTTPException(status_code=400, detail="项目下没有模块")
    
    # 获取所有模块的需求点
    requirement_points = db.query(RequirementPoint).filter(
        RequirementPoint.module_id.in_(module_ids)
    ).all()
    rp_ids = [rp.id for rp in requirement_points]
    rp_module_map = {rp.id: rp.module_id for rp in requirement_points}
    

    
    # 获取所有测试点
    test_points = db.query(TestPoint).filter(
        TestPoint.requirement_point_id.in_(rp_ids)
    ).all()
    tp_ids = [tp.id for tp in test_points]
    tp_rp_map = {tp.id: tp.requirement_point_id for tp in test_points}
    

    
    # 构建测试用例查询
    query = db.query(TestCase).filter(
        or_(
            TestCase.test_point_id.in_(tp_ids),
            TestCase.module_id.in_(module_ids),
            TestCase.project_id == project_id
        )
    )
    
    # 如果指定了ID，只导出指定的用例
    if request.ids:
        query = query.filter(TestCase.id.in_(request.ids))
    
    test_cases = query.order_by(TestCase.id).all()
    
    if not test_cases:
        raise HTTPException(status_code=400, detail="没有可导出的测试用例")
    
    # 根据格式生成不同的文件
    if request.format == "xmind":
        return export_to_xmind(project, modules, test_cases, tp_rp_map, rp_module_map, db)
    else:
        return export_to_excel(project, modules, test_cases, tp_rp_map, rp_module_map, module_map, db)


def export_to_excel(project, modules, test_cases, tp_rp_map, rp_module_map, module_map, db):
    """导出到Excel"""
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from io import BytesIO
    from datetime import datetime
    from urllib.parse import quote
    
    print(f"📊 开始生成Excel: 项目={project.name}, 用例数量={len(test_cases)}")
    
    # 移除语法错误的try语句，直接执行
    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"
    print(f"✅ 创建工作簿成功")
    
    # 定义样式
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 定义表头
    headers = ["序号", "所属模块", "用例标题", "前置条件", "测试步骤", "预期结果", "优先级", "测试分类", "设计方法", "状态"]
    
    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 优先级映射
    priority_map = {"high": "高", "medium": "中", "low": "低"}
    
    # 状态映射
    status_map = {"draft": "草稿", "under_review": "评审中", "approved": "已通过"}
    
    # 从数据库获取测试分类映射
    from app.models.settings import TestCategory
    categories = db.query(TestCategory).filter(TestCategory.is_active == True).all()
    category_map = {c.code: c.name for c in categories}
    
    # 从数据库获取设计方法映射
    from app.models.settings import TestDesignMethod
    methods = db.query(TestDesignMethod).filter(TestDesignMethod.is_active == True).all()
    method_map = {m.code: m.name for m in methods}
    
    # 写入数据
    for idx, tc in enumerate(test_cases, 1):
        row = idx + 1
        
        # 获取模块名称
        # 获取模块名称
        tp_id = tc.test_point_id
        rp_id = tp_rp_map.get(tp_id) if tp_id else None
        module_id = rp_module_map.get(rp_id) if rp_id else tc.module_id
        
        module_name = "未分类"
        if module_id:
            module_name = module_map.get(module_id, "未分类")
        elif tc.import_module_name:
            module_name = tc.import_module_name
        
        # 格式化测试步骤和预期结果
        steps_text = ""
        expected_text = ""
        if tc.test_steps and isinstance(tc.test_steps, list):
            steps_lines = []
            expected_lines = []
            for i, step in enumerate(tc.test_steps, 1):
                action = step.get("action", "") if isinstance(step, dict) else str(step)
                expected = step.get("expected", "") if isinstance(step, dict) else ""
                steps_lines.append(f"{i}. {action}")
                expected_lines.append(f"{i}. {expected}")
            steps_text = "\n".join(steps_lines)
            expected_text = "\n".join(expected_lines)
        
        # 写入单元格
        data = [
            idx,                                                    # 序号
            module_name,                                            # 模块名称
            tc.title or "",                                         # 用例标题
            tc.preconditions or "",                                 # 前置条件
            steps_text,                                             # 测试步骤
            expected_text,                                          # 预期结果
            priority_map.get(tc.priority, tc.priority or ""),       # 优先级
            category_map.get(tc.test_category, tc.test_category or ""),  # 测试分类
            method_map.get(tc.design_method, tc.design_method or ""),    # 设计方法
            status_map.get(tc.status, tc.status or "")              # 状态
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border
    
    # 调整列宽
    column_widths = [8, 15, 40, 25, 50, 50, 10, 12, 15, 10]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
    
    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 生成文件名
    filename = f"{project.name}_测试用例_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    encoded_filename = quote(filename)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
        }
    )


def export_to_xmind(project, modules, test_cases, tp_rp_map, rp_module_map, db):
    """
    导出到XMind思维导图 (XMind ZEN/2020+ JSON格式)
    """
    from fastapi.responses import StreamingResponse
    from datetime import datetime
    from urllib.parse import quote
    from app.utils.xmind_builder import XMindBuilder
    
    builder = XMindBuilder()
    root_topic = builder.set_root(f"{project.name} - 测试用例", "org.xmind.ui.map.logic.right")
    
    # 按模块分组用例
    module_cases = {}
    for tc in test_cases:
        tp_id = tc.test_point_id
        rp_id = tp_rp_map.get(tp_id)
        module_id = rp_module_map.get(rp_id) if rp_id else 0
        if module_id not in module_cases:
            module_cases[module_id] = []
        module_cases[module_id].append(tc)
    
    # 优先级标记
    priority_markers = {"high": "[高]", "medium": "[中]", "low": "[低]"}
    
    # 添加模块节点
    for module in modules:
        if module.id not in module_cases:
            continue
        
        module_topic = builder.create_topic(f"[模块] {module.name}")
        builder.add_child(root_topic, module_topic)
        
        # 添加用例节点
        for tc in module_cases[module.id]:
            priority_mark = priority_markers.get(tc.priority, "")
            case_topic = builder.create_topic(f"{priority_mark} {tc.title}")
            builder.add_child(module_topic, case_topic)
            
            # 前置条件
            if tc.preconditions:
                pre_topic = builder.create_topic(f"[前置条件] {tc.preconditions}")
                builder.add_child(case_topic, pre_topic)
            
            # 测试步骤
            if tc.test_steps and isinstance(tc.test_steps, list):
                steps_topic = builder.create_topic("[测试步骤]")
                builder.add_child(case_topic, steps_topic)
                
                expected_topic = builder.create_topic("[预期结果]")
                builder.add_child(case_topic, expected_topic)
                
                for i, step in enumerate(tc.test_steps, 1):
                    action = step.get("action", "") if isinstance(step, dict) else str(step)
                    expected = step.get("expected", "") if isinstance(step, dict) else ""
                    
                    if action:
                        step_topic = builder.create_topic(f"{i}. {action}")
                        builder.add_child(steps_topic, step_topic)
                    
                    if expected:
                        exp_topic = builder.create_topic(f"{i}. {expected}")
                        builder.add_child(expected_topic, exp_topic)
    
    # 处理未分类的用例
    if 0 in module_cases:
        other_topic = builder.create_topic("[模块] 未分类")
        builder.add_child(root_topic, other_topic)
        for tc in module_cases[0]:
            priority_mark = priority_markers.get(tc.priority, "")
            case_topic = builder.create_topic(f"{priority_mark} {tc.title}")
            builder.add_child(other_topic, case_topic)
    
    output = builder.build()
    
    # 生成文件名
    filename = f"{project.name}_测试用例_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xmind"
    encoded_filename = quote(filename)
    
    return StreamingResponse(
        output,
        media_type="application/octet-stream",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
        }
    )


@router.get("/projects/{project_id}/test-cases/template")
async def download_import_template(
    project_id: int,
    current_user: User = Depends(get_current_active_user)
):
    """下载测试用例导入模板"""
    import pandas as pd
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    from urllib.parse import quote
    
    # 定义表头
    headers = [
        "所属模块", "用例标题", "前置条件", "测试步骤", "预期结果", 
        "优先级(高/中/低)", "设计方法", "测试分类"
    ]
    
    # 创建示例数据
    example_data = [
        {
            "所属模块": "用户管理",
            "用例标题": "用户登录成功",
            "前置条件": "用户已注册且状态正常",
            "测试步骤": "1. 输入正确的用户名\n2. 输入正确的密码\n3. 点击登录按钮",
            "预期结果": "1. 登录成功\n2. 跳转至首页",
            "优先级(高/中/低)": "高",
            "设计方法": "功能测试",
            "测试分类": "功能测试"
        }
    ]
    
    df = pd.DataFrame(example_data, columns=headers)
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='导入模板')
        
        # 调整列宽
        worksheet = writer.sheets['导入模板']
        for i, col in enumerate(df.columns):
            column_len = max(df[col].astype(str).map(len).max(), len(col)) + 4
            worksheet.column_dimensions[chr(65 + i)].width = column_len
            
    output.seek(0)
    filename = "测试用例导入模板.xlsx"
    encoded_filename = quote(filename)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
        }
    )


@router.post("/projects/{project_id}/test-cases/import")
async def import_test_cases(
    project_id: int,
    file: UploadFile,
    auto_optimize: bool = Form(False),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """从Excel导入测试用例"""
    import pandas as pd
    from io import BytesIO
    from app.models.module import Module
    
    # 检查权限
    check_project_edit_permission(project_id, current_user, db)
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="仅支持 Excel 文件 (.xlsx, .xls)")
    
    try:
        # 读取Excel
        content = await file.read()
        df = pd.read_excel(BytesIO(content))
        
        # 验证表头
        required_columns = ["所属模块", "用例标题"]
        if not all(col in df.columns for col in required_columns):
            raise HTTPException(status_code=400, detail=f"文件缺少必要列: {', '.join(required_columns)}")
        
        # 预加载现有模块
        existing_modules = db.query(Module).filter(Module.project_id == project_id).all()
        module_map = {m.name: m.id for m in existing_modules}
        
        imported_count = 0
        new_cases = []
        
        for _, row in df.iterrows():
            title = str(row.get("用例标题", "")).strip()
            if not title or title == "nan":
                continue
                
            # 处理模块
            module_name = str(row.get("所属模块", "")).strip()
            if not module_name or module_name == "nan":
                module_name = "未分类"
                
            module_id = None
            import_module_name = None
            
            if module_name in module_map:
                module_id = module_map[module_name]
            else:
                # 不创建新模块，而是记录导入时的模块名称
                import_module_name = module_name
                # module_id 保持为 None，即归入"未分类"
            
            # 处理其他字段
            priority_map = {"高": "high", "中": "medium", "低": "low"}
            priority_raw = str(row.get("优先级(高/中/低)", "中")).strip()
            priority = priority_map.get(priority_raw, "medium")
            
            steps_raw = row.get("测试步骤", "")
            expected_raw = row.get("预期结果", "")
            
            # 解析步骤：按编号分割（支持 "1. xxx" 格式）
            test_steps = []
            if steps_raw and str(steps_raw) != "nan":
                import re
                steps_text = str(steps_raw)
                expected_text = str(expected_raw) if str(expected_raw) != "nan" else ""
                
                # 使用正则按 "数字." 分割
                step_pattern = re.compile(r'(\d+)\.\s*')
                step_parts = step_pattern.split(steps_text)
                expected_parts = step_pattern.split(expected_text)
                
                # step_parts: ['', '1', 'step1内容', '2', 'step2内容', ...]
                step_dict = {}
                for i in range(1, len(step_parts) - 1, 2):
                    num = step_parts[i]
                    content = step_parts[i + 1].strip() if i + 1 < len(step_parts) else ""
                    step_dict[num] = content
                
                expected_dict = {}
                for i in range(1, len(expected_parts) - 1, 2):
                    num = expected_parts[i]
                    content = expected_parts[i + 1].strip() if i + 1 < len(expected_parts) else ""
                    expected_dict[num] = content
                
                # 合并为 test_steps 列表
                if step_dict:
                    for num in sorted(step_dict.keys(), key=lambda x: int(x)):
                        test_steps.append({
                            "action": step_dict.get(num, ""),
                            "expected": expected_dict.get(num, "")
                        })
                else:
                    # 无法解析时，作为单个步骤
                    test_steps = [{"action": steps_text.strip(), "expected": expected_text.strip()}]
            
            new_case = TestCase(
                title=title,
                module_id=module_id,
                import_module_name=import_module_name,
                project_id=project_id,  # 直接关联项目
                preconditions=str(row.get("前置条件", "")) if str(row.get("前置条件", "")) != "nan" else None,
                test_steps=test_steps,
                expected_result=str(expected_raw) if str(expected_raw) != "nan" else None,
                priority=priority,
                design_method=str(row.get("设计方法", "")) if str(row.get("设计方法", "")) != "nan" else None,
                test_category=str(row.get("测试分类", "")) if str(row.get("测试分类", "")) != "nan" else None,
                created_by=current_user.id,
                status=TestCaseStatus.DRAFT,
                created_by_ai=False
            )
            db.add(new_case)
            imported_count += 1
            new_cases.append(new_case)
            
        db.commit()
        
        # 如果开启了自动优化 (这里仅做标记，实际优化逻辑可能需要异步任务)
        # TODO: Implement auto-optimize logic (call AI service)
        
        return {"success": True, "imported_count": imported_count, "message": f"成功导入 {imported_count} 条测试用例"}
        
    except Exception as e:
        db.rollback()
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"导入失败: {str(e)}")

